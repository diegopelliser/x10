#!/usr/bin/env python3
"""
Importa os dados das abas mensais do Excel para o Supabase.
Uso: python3 scripts/import_excel.py

Dependências: pip install openpyxl supabase python-dotenv
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
    from supabase import create_client
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Dependência ausente: {e}")
    print("Execute: pip install openpyxl supabase python-dotenv")
    sys.exit(1)

# ─── Configuração ──────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
load_dotenv(ROOT / ".env.local")

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
# Prefere a service_role key (bypassa RLS) — se não tiver, usa anon
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")
EXCEL_PATH       = ROOT / "Assets" / "Fluxo Financeiro.xlsx"
EXCEL_ABRIL_PATH = ROOT / "Assets" / "Dados Abril 2026.xlsx"
EXCEL_MAIO_PATH  = ROOT / "Assets" / "Fluxo Financeiro maio26.xlsx"

ABAS_MENSAIS = {
    "Setembro":       (9,  2025),
    "Outubro":        (10, 2025),
    "Novembro":       (11, 2025),
    "DEZEMBRO 2025":  (12, 2025),
    "JANEIRO 2026":   (1,  2026),
    "FEVEREIRO 2026": (2,  2026),
    "MARÇO 2026":     (3,  2026),
}

ABAS_ABRIL = {
    "ABRIL 2026": (4, 2026),
}

ABAS_MAIO = {
    "MAIO 2026": (5, 2026),
}

# Unidades que mapeiam para GDI (Garibaldi) — inclui verticais do escritório GDI
UNIDADES_GDI = [
    "x10 assessoria", "x10 assessor", "robôs quantum", "robos quantum",
    "quantum", "garibaldi", "gdi", "x10 ai inv", "ai inv gdi",
    "consórcio", "consorcio", "corban", "vida e previdência", "vida e previdencia",
    "imobiliário", "imobiliario",
]

# Unidades que mapeiam para POA
UNIDADES_POA = [
    "poa", "porto alegre", "ai inv poa", "x10 poa",
]

# Prefixos de linhas de resumo/totalizadores no col-A (devem ser ignoradas como grupo)
PREFIXOS_RESUMO = (
    "faturamento", "despesas", "resultado", "total geral",
)

# Categorias da planilha que são impostos
CATS_IMPOSTO = [
    "inss", "iss", "simples", "das", "fgts", "irrf", "pis", "cofins",
    "imposto", "tributo", "gps", "darf", "imposto retido", "impostos retidos",
]

# Categorias da planilha que são custos variáveis
CATS_CUSTO = [
    "comissões de vendedores", "comissao de vendedor", "comissão de vendedor",
    "remuneração estagiário", "remuneracao estagiario",
]

# Categorias que são receita (quando valor positivo)
CATS_RECEITA = [
    "receitas de serviços", "receita de serviço", "rendimentos de aplicações",
    "rendimentos financeiros", "premiações", "premiacao", "prêmios",
    "descontos incondicionais", "outras receitas",
]


def normalizar(texto: str) -> str:
    return texto.lower().strip() if texto else ""


def inferir_unidade(raw: str) -> str:
    n = normalizar(raw)
    if not n:
        return "OUTROS"
    for chave in UNIDADES_GDI:
        if chave in n:
            return "GDI"
    for chave in UNIDADES_POA:
        if chave in n:
            return "POA"
    return "OUTROS"


def inferir_tipo_negativo(categoria: str) -> str:
    """Infere tipo para valores negativos (despesas/custos/impostos).
    Custo é checado antes de imposto para evitar falso-positivo:
    'iss' é substring de 'comissões', o que causaria classificação errada."""
    n = normalizar(categoria)
    for p in CATS_CUSTO:
        if p in n:
            return "custo"
    for p in CATS_IMPOSTO:
        if p in n:
            return "imposto"
    return "despesa"


def inferir_categoria(categoria_raw: str, tipo: str) -> str:
    """Normaliza categoria bruta para as categorias padrão do sistema."""
    n = normalizar(categoria_raw)

    if tipo == "receita":
        if "rendimento" in n or "aplicaç" in n:
            return "Rendimentos Financeiros"
        if "premiaç" in n or "prêmio" in n or "premio" in n:
            return "Outras Receitas"
        if "desconto incondicional" in n or "restituiç" in n:
            return "Outras Receitas"
        # Receita genérica — será refinada pela unidade no futuro
        return "Outras Receitas"

    if tipo == "imposto":
        if "inss" in n or "gps" in n:
            return "INSS sobre Salários"
        if "iss" in n:
            return "ISS sobre Faturamento"
        if "simples" in n or "das" in n:
            return "Simples Nacional (DAS)"
        if "fgts" in n:
            return "FGTS"
        if "irrf" in n or "retid" in n:
            return "Impostos Retidos"
        return "Outros Impostos"

    if tipo == "custo":
        # "comiss" cobre singular (comissão) e plural (comissões) — caracteres Unicode distintos
        if "comiss" in n:
            return "Comissões de Vendedores"
        # "estagi" cobre estagiário e estagiários
        if "estagi" in n:
            return "Remuneração de Estagiários"
        return "Outros Custos Variáveis"

    # despesa
    if "salário" in n or "salario" in n or "folha" in n:
        return "Folha de Pagamento"
    if "pró-labore" in n or "prolabore" in n or "pro-labore" in n or "pro labore" in n:
        return "Pró-labore"
    if "13" in n or "férias" in n or "ferias" in n:
        return "13º Salário / Férias"
    if "encargo" in n or "previdência patronal" in n:
        return "Encargos Sociais"
    if "aluguel" in n:
        return "Aluguel"
    if "energia" in n or "internet" in n or "condomínio" in n or "condominio" in n or "luz" in n:
        return "Energia / Internet / Condomínio"
    if "software" in n or "licença" in n or "licenca" in n or "sistema" in n or "tecnologia" in n:
        return "Tecnologia e Software"
    if "contábil" in n or "contabil" in n or "contabilidade" in n:
        return "Honorários Contábeis"
    if "honorário" in n or "honorario" in n:
        return "Honorários Contábeis"
    if "jurídic" in n or "juridic" in n or "advocaci" in n:
        return "Honorários Jurídicos"
    if "marketing" in n or "evento" in n or "brind" in n or "publicidade" in n:
        return "Marketing e Eventos"
    if "material" in n or "bens" in n or "equipamento" in n:
        return "Bens e Materiais"
    if "tarifa" in n or "bancária" in n or "bancaria" in n:
        return "Outras Despesas"
    if "exame" in n or "médico" in n or "medico" in n or "saúde" in n:
        return "Outras Despesas"
    if "curso" in n or "treinamento" in n or "capacitaç" in n:
        return "Outras Despesas"
    if "devolução" in n or "devolucao" in n or "cancelamento" in n:
        return "Outras Despesas"

    return "Outras Despesas"


def inferir_receita_por_descricao(descricao: str, unidade_raw: str) -> str:
    """Mapeia receitas para categorias baseado na descrição e unidade."""
    n  = normalizar(descricao)
    un = normalizar(unidade_raw)

    if "quantum" in un or "robô" in un:
        return "Assessoria de Investimentos"
    if "corban" in un or "corban" in n:
        return "Comissões Corban"
    if "consórcio" in un or "consorcio" in un or "consórcio" in n or "rodobens" in n or "newcom" in n:
        return "Comissões Consórcios"
    if "v&p" in un or "vida" in un or "previdência" in un or "v&p" in n or "btg pactual corretora" in n:
        return "Comissões V&P"
    if "imobiliário" in un or "imobiliario" in un or "captal" in n:
        return "Outras Receitas"
    if "assessoria" in un or "x10 assessor" in un or "btg pactual" in n:
        return "Assessoria de Investimentos"
    if "poa" in un or "ai inv poa" in un:
        return "Assessoria de Investimentos"
    return "Outras Receitas"


def eh_linha_resumo(unidade_raw: str) -> bool:
    """Retorna True para linhas de totalizadores (Faturamento X, Despesas X, Resultado X)."""
    n = unidade_raw.lower().strip()
    return any(n.startswith(p) for p in PREFIXOS_RESUMO)


def ler_aba(ws, mes: int, ano: int) -> tuple[list[dict], list[dict]]:
    registros = []
    ignorados = []

    # Carry-forward: unidade, centro_custo e categoria só aparecem na 1ª linha do grupo
    unidade_atual      = ""
    centro_custo_atual = ""
    categoria_atual    = ""

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        unidade_raw  = str(row[0]).strip() if row[0] else ""
        categoria    = str(row[1]).strip() if row[1] else ""
        descricao    = str(row[2]).strip() if row[2] else ""
        valor_raw    = row[3]  # Coluna D = Valor

        # Ignorar linhas de TOTAL ou completamente vazias
        if unidade_raw.upper() == "TOTAL":
            unidade_atual      = ""
            centro_custo_atual = ""
            categoria_atual    = ""
            continue
        if not any([unidade_raw, categoria, descricao, valor_raw]):
            continue

        # Ignorar linhas de consolidação/totalização do Excel
        desc_upper = descricao.upper()
        if any(p in desc_upper for p in (
            "TOTAL", "SALDO FINAL", "SALDO CONSORCIOS", "SALDO CORBAN",
            "SALDO IMOBILI", "SALDO POA", "SALDO QUANTUM", "SALDO V&P",
            "SALDO X10", "SALDO GDI", "RESULTADO FINAL",
        )):
            continue

        # Linhas de resumo/totalização (Faturamento X, Resultado X, Total geral, etc.)
        # NÃO atualizam o carry-forward E são completamente ignoradas — não são lançamentos reais
        if unidade_raw and eh_linha_resumo(unidade_raw):
            continue

        if unidade_raw:
            unidade_atual      = unidade_raw
            centro_custo_atual = unidade_raw  # guarda o nome do vertical como centro de custo
        if categoria:
            categoria_atual = categoria

        cat_efetiva = categoria_atual

        # Ignorar linhas de cabeçalho da planilha
        if cat_efetiva.lower() in ("categoria", "descrição", "nan", "none", ""):
            continue

        # Processar valor
        if valor_raw is None:
            ignorados.append({"linha": i, "motivo": "valor vazio", "categoria": cat_efetiva, "descricao": descricao})
            continue

        try:
            valor_num = float(str(valor_raw).replace(",", ".").replace("R$", "").strip())
        except (ValueError, TypeError):
            ignorados.append({"linha": i, "motivo": f"valor não numérico ({valor_raw!r})", "categoria": cat_efetiva, "descricao": descricao})
            continue

        if valor_num == 0:
            continue

        valor_abs = abs(valor_num)
        unidade   = inferir_unidade(unidade_atual)

        if valor_num > 0:
            # Receita
            tipo    = "receita"
            cat_std = inferir_receita_por_descricao(descricao, unidade_atual)
        else:
            # Despesa / custo / imposto
            tipo    = inferir_tipo_negativo(cat_efetiva)
            cat_std = inferir_categoria(cat_efetiva, tipo)

        registros.append({
            "mes":          mes,
            "ano":          ano,
            "unidade":      unidade,
            "tipo":         tipo,
            "categoria":    cat_std,
            "descricao":    descricao or cat_efetiva,
            "valor":        round(valor_abs, 2),
            "centro_custo": centro_custo_atual or None,
        })

    return registros, ignorados


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("Erro: NEXT_PUBLIC_SUPABASE_URL e NEXT_PUBLIC_SUPABASE_ANON_KEY precisam estar em .env.local")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"Erro: arquivo não encontrado: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Abrindo {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    total_importados = 0
    todos_ignorados: list[dict] = []

    for nome_aba, (mes, ano) in ABAS_MENSAIS.items():
        if nome_aba not in wb.sheetnames:
            print(f"  ⚠  Aba '{nome_aba}' não encontrada — pulando")
            continue

        ws = wb[nome_aba]
        registros, ignorados = ler_aba(ws, mes, ano)

        for item in ignorados:
            todos_ignorados.append({**item, "aba": nome_aba})

        if not registros:
            print(f"  ⚠  '{nome_aba}': nenhum registro válido ({len(ignorados)} ignorados)")
            continue

        # Limpar dados existentes do período antes de inserir
        client.table("lancamentos").delete().eq("mes", mes).eq("ano", ano).execute()

        # Inserir em lotes de 100
        for j in range(0, len(registros), 100):
            lote = registros[j:j+100]
            client.table("lancamentos").insert(lote).execute()

        print(f"  ✓  '{nome_aba}': {len(registros)} importados, {len(ignorados)} ignorados")
        total_importados += len(registros)

    # ── Abril 2026 (arquivo separado) ──────────────────────────
    if EXCEL_ABRIL_PATH.exists():
        print(f"\nAbrindo {EXCEL_ABRIL_PATH.name}...")
        wb_abril = openpyxl.load_workbook(EXCEL_ABRIL_PATH, data_only=True)
        for nome_aba, (mes, ano) in ABAS_ABRIL.items():
            if nome_aba not in wb_abril.sheetnames:
                print(f"  ⚠  Aba '{nome_aba}' não encontrada em {EXCEL_ABRIL_PATH.name} — pulando")
                continue
            ws = wb_abril[nome_aba]
            registros, ignorados = ler_aba(ws, mes, ano)
            for item in ignorados:
                todos_ignorados.append({**item, "aba": nome_aba})
            if not registros:
                print(f"  ⚠  '{nome_aba}': nenhum registro válido")
                continue
            client.table("lancamentos").delete().eq("mes", mes).eq("ano", ano).execute()
            for j in range(0, len(registros), 100):
                client.table("lancamentos").insert(registros[j:j+100]).execute()
            print(f"  ✓  '{nome_aba}': {len(registros)} importados, {len(ignorados)} ignorados")
            total_importados += len(registros)

    # ── Maio 2026 ──────────────────────────────────────────────────
    if EXCEL_MAIO_PATH.exists():
        print(f"\nAbrindo {EXCEL_MAIO_PATH.name}...")
        wb_maio = openpyxl.load_workbook(EXCEL_MAIO_PATH, data_only=True)
        for nome_aba, (mes, ano) in ABAS_MAIO.items():
            if nome_aba not in wb_maio.sheetnames:
                print(f"  ⚠  Aba '{nome_aba}' não encontrada em {EXCEL_MAIO_PATH.name} — pulando")
                continue
            ws = wb_maio[nome_aba]
            registros, ignorados = ler_aba(ws, mes, ano)
            for item in ignorados:
                todos_ignorados.append({**item, "aba": nome_aba})
            if not registros:
                print(f"  ⚠  '{nome_aba}': nenhum registro válido")
                continue
            client.table("lancamentos").delete().eq("mes", mes).eq("ano", ano).execute()
            for j in range(0, len(registros), 100):
                client.table("lancamentos").insert(registros[j:j+100]).execute()
            print(f"  ✓  '{nome_aba}': {len(registros)} importados, {len(ignorados)} ignorados")
            total_importados += len(registros)

    print(f"\nConcluído: {total_importados} registros importados, {len(todos_ignorados)} linhas ignoradas.")

    if todos_ignorados:
        print("\n── Linhas ignoradas ──────────────────────────────────────")
        for item in todos_ignorados:
            print(f"  [{item['aba']}] linha {item['linha']:>3}  {item['motivo']}")
            print(f"           categoria={item['categoria']!r}  descricao={item['descricao']!r}")


if __name__ == "__main__":
    main()

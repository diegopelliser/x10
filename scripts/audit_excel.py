#!/usr/bin/env python3
"""
Audita o Excel sem tocar no banco — mostra totais por aba e lista o que seria importado.
Uso: python3 scripts/audit_excel.py
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Execute: pip install openpyxl")
    sys.exit(1)

ROOT       = Path(__file__).parent.parent
EXCEL_PATH = ROOT / "Assets" / "Fluxo Financeiro.xlsx"

ABAS_MENSAIS = {
    "Setembro":       (9,  2025),
    "Outubro":        (10, 2025),
    "Novembro":       (11, 2025),
    "DEZEMBRO 2025":  (12, 2025),
    "JANEIRO 2026":   (1,  2026),
    "FEVEREIRO 2026": (2,  2026),
    "MARÇO 2026":     (3,  2026),
}

def main():
    if not EXCEL_PATH.exists():
        print(f"Erro: arquivo não encontrado: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Arquivo: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print(f"\nAbas encontradas no Excel ({len(wb.sheetnames)}):")
    for nome in wb.sheetnames:
        marcador = "  ✓" if nome in ABAS_MENSAIS else "  —"
        print(f"{marcador}  '{nome}'")

    print("\n" + "─" * 60)
    print(f"{'ABA':<20} {'LINHAS':>6}  {'VÁLIDAS':>7}  {'IGNORADAS':>9}  {'TOTAL R$':>14}")
    print("─" * 60)

    grand_total = 0.0
    for nome_aba, (mes, ano) in ABAS_MENSAIS.items():
        if nome_aba not in wb.sheetnames:
            print(f"{'  ⚠ ' + nome_aba:<20}  NÃO ENCONTRADA")
            continue

        ws = wb[nome_aba]
        validas = 0
        ignoradas = 0
        total = 0.0
        linhas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            linhas += 1
            valor_raw = row[3]
            categoria = str(row[1]).strip() if row[1] else ""

            if not categoria or categoria.lower() in ("categoria", "descrição", "nan", "none", ""):
                ignoradas += 1
                continue

            try:
                valor = float(str(valor_raw).replace(",", ".").replace("R$", "").strip())
                if valor <= 0:
                    ignoradas += 1
                    continue
                validas += 1
                total += valor
            except (ValueError, TypeError):
                ignoradas += 1

        grand_total += total
        print(f"  {nome_aba:<18} {linhas:>6}  {validas:>7}  {ignoradas:>9}  {total:>14,.2f}")

    print("─" * 60)
    print(f"  {'TOTAL GERAL':<18} {'':>6}  {'':>7}  {'':>9}  {grand_total:>14,.2f}")

    # Detalhe da primeira aba: todas as colunas de cada linha não-vazia (para entender a estrutura)
    primeira_aba = next(iter(ABAS_MENSAIS))
    if primeira_aba in wb.sheetnames:
        ws = wb[primeira_aba]
        print(f"\n\n{'═' * 80}")
        print(f"ESTRUTURA RAW — aba '{primeira_aba}' (primeiras 30 linhas não-vazias, todas as colunas)")
        print(f"{'═' * 80}")
        print(f"  {'LN':>3}  {'COL-A':<20}  {'COL-B':<30}  {'COL-C':<25}  {'COL-D':>12}  {'COL-E':>12}  {'COL-F':>12}")
        count = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not any(row):
                continue
            cols = [str(c or "").strip()[:25] for c in (list(row) + [""] * 6)[:6]]
            print(f"  {i:>3}  {cols[0]:<20}  {cols[1]:<30}  {cols[2]:<25}  {cols[3]:>12}  {cols[4]:>12}  {cols[5]:>12}")
            count += 1
            if count >= 30:
                print("  ...")
                break

if __name__ == "__main__":
    main()
